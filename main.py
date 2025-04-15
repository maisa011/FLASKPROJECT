import csv
from datetime import datetime

from flask import send_file, redirect, url_for
from io import BytesIO
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows



from flask import Flask, render_template, request, send_file, flash
from io import BytesIO

from markupsafe import Markup
import plotly.graph_objects as go
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook



app = Flask(__name__)
output_excel = BytesIO()  # გლობალური მეხსიერების ბუფერი



@app.route('/')
def home():
    return render_template("index.html")

@app.route('/PspToAversi')
def user():
    return render_template("PspToAversi.html")

@app.route('/interactive')
def interactive_chart():


    return render_template("interactive.html")


@app.route('/sales')
def sales():
    return render_template('sales.html')



@app.route('/weather', methods=['GET', 'POST'])
def weather():
    global output_excel  # ვიყენებთ გლობალურ ბუფერს გადმოსაწერად

    if request.method == 'POST':
        try:
            # ფაილების ატვირთვა
            medicine_file = request.files.get('medicine')
            erp_file = request.files.get('erp')
            aramedi_file = request.files.get('aramedi')

            # DataFrame-ებად გარდაქმნა
            df1 = pd.read_excel(medicine_file, dtype={"ფარმაცევტის პირადი N": str}) if medicine_file else pd.DataFrame()
            df2 = pd.read_excel(erp_file, dtype={"პირადი ნომერი": str}) if erp_file else pd.DataFrame()
            df3 = pd.read_excel(aramedi_file, dtype={"ფარმაცევტის პირადი N": str}) if aramedi_file else pd.DataFrame()

            # დაამუშავე
            final_df = bonusFormireba(df1, df2, df3)

            if final_df is None or final_df.empty:
                return render_template("weather.html", message="❌ ფაილი არ დამუშავდა. გადაამოწმეთ შეყვანილი ფაილები.")

            # Excel ფაილის გენერაცია მეხსიერებაში
            output_excel = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            for row in dataframe_to_rows(final_df, index=False, header=True):
                ws.append(row)

            wb.save(output_excel)
            output_excel.seek(0)

            return render_template("weather.html", table=final_df.to_html(classes="table", border=1))

        except Exception as e:
            print(f"შეცდომა weather-ში: {e}")
            return render_template("weather.html", message="⚠️ დაფიქსირდა შეცდომა. გთხოვთ გადაამოწმოთ ფაილები.")

    return render_template("weather.html")


@app.route('/bonus',methods = ['GET', 'POST'] )
def bonuses():
    return render_template("weather.html")

@app.route('/download_bonus')
def download_bonus():
    global output_excel
    output_excel.seek(0)  # დავაბრუნოთ პოზიცია ფაილის დასაწყისში
    return send_file(output_excel, download_name="bonus.xlsx", as_attachment=True)

@app.route('/convert_psp', methods=['POST'])
def convert_psp():
    excel_file = request.files.get('excel_file')

    if not excel_file:
        return "გთხოვ აირჩიო ფაილი", 400

    try:
        # ატვირთული PSP ფაილის წაკითხვა
        df = pd.read_excel(excel_file)
        df = df.fillna("")

        # data.csv ფაილის წაკითხვა (უნდა იდოს აპლიკაციის ფოლდერში)
        data = pd.read_csv("data.csv")

        # გაერთიანება
        merged = pd.merge(df, data, left_on="kod", right_on="კოდი psp", how='left')

        if "არტიკული" in merged.columns:
            merged["არტიკული"] = merged["არტიკული"].fillna("").apply(
                lambda x: str(int(x)) if str(x).replace('.', '', 1).isdigit() else "")

        if "ქარხ.კოდი" in merged.columns:
            merged["ქარხ.კოდი"] = merged["ქარხ.კოდი"].fillna("").apply(
                lambda x: str(int(x)) if str(x).replace('.', '', 1).isdigit() else "")

        merged = merged.drop(columns=["კოდი psp", "მედიკამენტი"], errors='ignore')

        # reorder
        columns = ["არტიკული", "საქონლის დასახელება"] + [col for col in merged.columns if col not in ["არტიკული", "საქონლის დასახელება"]]
        merged = merged[columns]

        # ჩაწერა, როცა არტიკული ცარიელია
        if "არტიკული" in merged.columns and "kod" in merged.columns and "საქონლის დასახელება" in merged.columns:
            merged.loc[merged["არტიკული"] == "", "საქონლის დასახელება"] = (
                "შესაბამისობა არ მოიძებნა პსპს კოდია: " + merged["kod"].astype(str)
            )

        merged = merged.drop(merged.columns[[2, 3]], axis=1)
        columns = list(merged.columns)
        third_column = columns.pop(2)
        columns.append(third_column)
        merged = merged[columns]

        # არასაჭირო სვეტების წაშლა
        cols_to_drop = ["name_manuf", "kod_manuf", "bar", "stran", "country", "root_id", "p_nalog",
                        "kod_pol", "address", "n_zed", "Unnamed: 0", "ქარხ.კოდი", "kod"]
        merged.drop(columns=[col for col in cols_to_drop if col in merged.columns], inplace=True)

        if "name_geo" in merged.columns:
            merged.rename(columns={"name_geo": "psp დასახელება"}, inplace=True)

        # Excel-ში ჩასაწერად ფაილის შექმნა მეხსიერებაში
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        for row in dataframe_to_rows(merged, index=False, header=True):
            ws.append(row)

        # ავტომატური სვეტის სიგრძე
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        # ყვითელი ფონი შესაბამის უჯრებში
        if "საქონლის დასახელება" in merged.columns:
            col_index = merged.columns.get_loc("საქონლის დასახელება") + 1
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index, max_row=ws.max_row):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("შესაბამისობა არ მოიძებნა პსპს კოდია"):
                        cell.fill = yellow_fill

        wb.save(output)
        output.seek(0)

        return send_file(output, download_name="კონვერტირებული_ფაილი.xlsx", as_attachment=True)

    except Exception as e:
        error_message = f"შეცდომა დამუშავებისას: {str(e)}"
        # შეცდომის შემთხვევაში წაიკითხავს psp.html-ს
        return render_template("psp.html", error=error_message)



@app.route('/convert_aversi', methods=['POST'])
def convert_aversi():
    file = request.files.get('excel_file')

    if not file:
        return "ფაილი არ არის არჩეული", 400

    try:
        df = pd.read_excel(file).fillna("")

        data = pd.read_csv("dataAversi.csv")

        # 'kod' ველის სტრინგად გარდაქმნა ორივე მხარეს
        df["kod"] = df["kod"].astype(str).str.strip().str.replace(".0", "", regex=False)
        data["კოდი psp"] = data["კოდი psp"].astype(str).str.strip()

        merged = pd.merge(df, data, left_on="kod", right_on="კოდი psp", how="left")

        # არტიკული და ქარხ.კოდი სტრინგად
        if "არტიკული" in merged.columns:
            merged["არტიკული"] = merged["არტიკული"].fillna("").apply(
                lambda x: str(int(x)) if str(x).replace('.', '', 1).isdigit() else "")

        if "ქარხ.კოდი" in merged.columns:
            merged["ქარხ.კოდი"] = merged["ქარხ.კოდი"].fillna("").apply(
                lambda x: str(int(x)) if str(x).replace('.', '', 1).isdigit() else "")

        # არ უნდა წაიშალოს 'kod' მანამ სანამ გამოიყენება
        if "არტიკული" in merged.columns and "საქონლის დასახელება" in merged.columns and "kod" in merged.columns:
            merged["საქონლის დასახელება"] = merged.apply(
                lambda row: f"შესაბამისობა არ მოიძებნა პსპს კოდია: {row['kod']}"
                if row["არტიკული"] == "" else row["საქონლის დასახელება"],
                axis=1
            )

        # არასაჭირო სვეტების წაშლა
        drop_cols = [
            "კოდი psp", "მედიკამენტი", "name_manuf", "kod_manuf", "bar", "stran", "country", "root_id",
            "p_nalog", "kod_pol", "address", "n_zed", "Unnamed: 0", "ქარხ.კოდი", "kod"
        ]
        for col in drop_cols:
            if col in merged.columns:
                merged.drop(columns=col, inplace=True)

        # სვეტების გადალაგება
        main_cols = ["არტიკული", "საქონლის დასახელება"]
        other_cols = [col for col in merged.columns if col not in main_cols]
        merged = merged[main_cols + other_cols]

        # გადარქმევა
        if "name_geo" in merged.columns:
            merged.rename(columns={"name_geo": "psp დასახელება"}, inplace=True)

        # თარიღის კონვერტაცია
        if "vada" in merged.columns:
            merged["vada"] = merged["vada"].apply(convert_date_format)

        # Excel ფაილის შექმნა
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        for row in dataframe_to_rows(merged, index=False, header=True):
            ws.append(row)

        # სვეტების სიგანის მორგება
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

        # ყვითელი ფერის მონიშვნა
        if "საქონლის დასახელება" in merged.columns:
            col_index = merged.columns.get_loc("საქონლის დასახელება") + 1
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index, max_row=ws.max_row):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("შესაბამისობა არ მოიძებნა"):
                        cell.fill = yellow_fill

        wb.save(output)
        output.seek(0)

        return send_file(output, download_name="ავერსი_დამუშავებული.xlsx", as_attachment=True)

    except Exception as e:
        error_message = f"შეცდომა დამუშავებისას: {str(e)}"
        return render_template("aversi.html", error=error_message)

@app.route('/psp')
def psp_page():
    return render_template("psp.html")

@app.route('/aversi')
def aversi_page():
    return render_template("aversi.html")



COMMENTS_FILE = 'comments.csv'
@app.route('/notes')
def notes():
    comments = []
    try:
        with open(COMMENTS_FILE, 'r', encoding='utf-8') as file:
            reader = csv.reader(file)
            for row in reader:
                if len(row) == 3:  # ✅ მხოლოდ მაშინ დაამატე, როცა 3 ელემენტია
                    timestamp, name, comment = row
                    comments.append((timestamp, name, comment))
    except FileNotFoundError:
        pass
    return render_template("notes.html", comments=comments)


@app.route('/submit_note', methods=['POST'])
def submit_note():
    name = request.form['name']
    comment = request.form['comment']
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with open(COMMENTS_FILE, 'a', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow([timestamp, name, comment])

    return redirect(url_for('notes'))




# თარიღის ფორმატის გარდაქმნა: "MM/YYYY" → "01.MM.YYYY"
def convert_date_format(date_str):
    try:
        month, year = date_str.split("/")
        return f"01.{month}.{year}"
    except Exception:
        return None


def bonusFormireba(df1,df2,df3):

        try:
            if not df1.empty and not df2.empty and not df3.empty:
                print("3 ფაილს ვერ აირჩევ ერთიანად")
                return pd.DataFrame


            if not df2.empty and not df1.empty:
                df2 = df2.drop(df2.index[0]).reset_index(drop=True)

                new_column_names = {
                    6: "რაჩენის გეგმა-ფაქტი",
                    7: "რაჩენის გეგმის შესრულების %",
                    8: "რაჩენის გასაცემი ბონუსი",
                    9: "ინდივიდუალური გეგმა",
                    10: "ინდივიდუალური ფაქტი",
                    11: "ინდივიდუალური გეგმის სესრულების %",
                    12: "თვითმომსახურების დამატებითი ბონუსი",
                    13: "თვითმომსახურების გაყიდვები(გეგმა)",
                    14: "თვითმომსახურების გაყიდვები(ფაქტი)",
                    15: "თვითმომსახურების გაყიდვები(შესრულების %)",
                    16: "თვითმომსახურების გაყიდვები(დამატებითი ბონუსი)",
                    17: "არამედების გეგმა",
                    18: "არამედების ფაქტი",
                    19: "არამედების შესრულების %",
                    20: "არამედების თვითმომსახურების დამატებითი ბონუსი",
                    21: "არამედების გაყიდვები ფაქტი",
                    22: "არამედების გაყიდვები არამედების 1% (გეგმის შესრ)",
                    23: "impo გასაცემი ბონუსი",
                    24: "local2 გასაცემი ბონუსი",
                    25: "LocalOmronTonus გასაცემი ბონუსი",
                    26: "Promo გასაცემი ბონუსი",
                    27: "სულ გასაცემი ბონუსი"
                }

                columns = list(df2.columns)
                for index, new_name in new_column_names.items():
                    if index < len(columns):
                        columns[index] = new_name
                df2.columns = columns

                required_columns = [18, 19, 22, 27]
                if not all(idx < len(df2.columns) for idx in required_columns):
                    print("საჭირო სვეტები არ არსებობს df2-ში.")



                for col_idx in required_columns:
                    df2.iloc[:, col_idx] = pd.to_numeric(df2.iloc[:, col_idx], errors='coerce').fillna(0)

                condition = df2.iloc[:, 19] >= 95
                df2.iloc[:, 22] = df2.iloc[:, 22].where(condition, 0.0)

                mapping = {
                    20: 0, 0: 1, 1: 2, 2: 4, 3: 5, 11: 6, 14: 7,
                    4: 8, 6: 9, 7: 10, 10: 11, 8: 21, 9: 22,
                    16: 23, 18: 24, 15: 25, 17: 26, 19: 27
                }

                df1.iloc[:, 19] = df1.iloc[:, 19] - df1.iloc[:, 9]

                mapped_df2 = pd.DataFrame(columns=df1.columns)
                for df1_idx, df2_idx in mapping.items():
                    if df1_idx < len(df1.columns) and df2_idx < len(df2.columns):
                        mapped_df2[df1.columns[df1_idx]] = df2.iloc[:, df2_idx]

                df1 = pd.concat([df1, mapped_df2], ignore_index=True)
                df1 = df1.fillna(0)
                df1 = df1[df1["ფარმაცევტის პირადი N"] != 0].reset_index(drop=True)

                if 20 < len(df1.columns):
                    df1.iloc[:, 20] = df1.iloc[:, 20].replace({
                        "თბილისის აფთიაქების საწყობები": "თბილისი",
                        "რეგიონის აფთიაქების საწყობები": "რეგიონი"
                    })

                if "ფარმაცევტის პირადი N" in df1.columns:
                    df1["ფარმაცევტის პირადი N"] = df1["ფარმაცევტის პირადი N"].str.lstrip('0')

                grouped_df = df1.groupby(["აფთიაქის კოდი", "ფარმაცევტის პირადი N"], as_index=False).agg({
                    "ფარმაცევტის სახელი, გვარი": "first",
                    "აფთიაქის დასახელება": "first",
                    "რაჩენის ჯილდო": "sum",
                    "რანდომის ჯილდო": "first",
                    "პირადი გეგმა": "first",
                    "პირადი რეალიზაცია": "sum",
                    "პირადი რეალიზაცია (არამედი)": "sum",
                    "არამედის რეალიზაციის ბონუსი": "sum",
                    "შესრულების %": "first",
                    "Rachen რეალიზაცია": "sum",
                    "Rachen1 რეალიზაცია": "first",
                    "Rachen2 რეალიზაცია": "first",
                    "რაჩენის შესრულება": "first",
                    "ლოკალი.ომრონი/ტონუსი": "sum",
                    "იმპორტი": "sum",
                    "პრომო": "sum",
                    "ლოკალ2": "sum",
                    "არამედ ბონუსი": "sum",
                    "Region": "first"
                })

                df1 = grouped_df
                df1.iloc[:, 19] = df1.iloc[:, 19] + df1.iloc[:, 9]
                numeric_cols = df1.columns[2:]
                df1[numeric_cols] = df1[numeric_cols].applymap(
                    lambda x: round(x, 2) if isinstance(x, (int, float)) else x)
                return df1

            if not df3.empty and not df2.empty:
                df2 = df2[df2["მომხმარებლების ჯგუფი"] == "არამედიკამენტის კონსულტანტი"]
                df3.iloc[:, 21] = df3.iloc[:, 21] - df3.iloc[:, 16]
                df2 = df2.reset_index(drop=True)
                condition = df2.iloc[:, 19] >= 95
                df2.iloc[:, 22] = df2.iloc[:, 22].where(condition, 0.0)

                mapping = {
                    0: 1, 1: 2, 2: 4, 3: 5, 4: 17, 5: 18, 6: 19,
                    7: 21, 8: 25, 10: 23, 12: 26, 14: 24, 16: 22,
                    21: 27, 22: 0, 23: None
                }

                mapped_df2 = pd.DataFrame(0, index=range(len(df2)), columns=df3.columns)

                for df3_idx, df2_idx in mapping.items():
                    if df2_idx is None:
                        continue
                    if df3_idx >= len(df3.columns) or df2_idx >= len(df2.columns):
                        print(f"მოსალოდნელი ინდექსი df3 ან df2 არ არსებობს: df3_idx={df3_idx}, df2_idx={df2_idx}")
                        continue
                    mapped_df2[df3.columns[df3_idx]] = df2.iloc[:, df2_idx]

                df3 = df3.fillna(0).infer_objects()
                df3.iloc[:, 21] = df3.iloc[:, 21] + df3.iloc[:, 16]
                df3 = pd.concat([df3, mapped_df2], ignore_index=True).fillna(0).infer_objects()

                if "Region" in df3.columns:
                    df3["Region"] = df3["Region"].replace({
                        "თბილისის აფთიაქების საწყობები": "თბილისი",
                        "რეგიონის აფთიაქების საწყობები": "რეგიონი"
                    })

                columns_to_convert = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21]
                for column in columns_to_convert:
                    if column < len(df3.columns):
                        df3.iloc[:, column] = pd.to_numeric(df3.iloc[:, column], errors='coerce').round(2)

                return df3
            else:
                return pd.DataFrame






        except Exception as e:
            print(f"შეცდომა bonusformireba-ში: {e}")




if __name__ == "__main__":
    app.run(debug=True)
